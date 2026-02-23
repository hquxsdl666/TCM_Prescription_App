"""
Excel导出模块
将处方数据导出为Excel文件
"""

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        self.headers = [
            '序号',
            '患者姓名',
            '年龄',
            '性别',
            '日期',
            '症状',
            '诊断',
            '方剂名称',
            '药材组成',
            '剂量',
            '用法',
            '医师',
            '医院',
            '备注'
        ]
    
    def export(self, prescriptions, output_path=None):
        """
        导出处方数据到Excel
        
        Args:
            prescriptions: 处方数据列表
            output_path: 输出文件路径
        
        Returns:
            输出文件路径
        """
        if output_path is None:
            # 默认文件名
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'prescriptions_{timestamp}.xlsx'
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '处方记录'
        
        # 设置标题行样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入标题行
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置列宽
        column_widths = {
            'A': 8,   # 序号
            'B': 12,  # 患者姓名
            'C': 8,   # 年龄
            'D': 8,   # 性别
            'E': 12,  # 日期
            'F': 30,  # 症状
            'G': 25,  # 诊断
            'H': 20,  # 方剂名称
            'I': 40,  # 药材组成
            'J': 10,  # 剂量
            'K': 25,  # 用法
            'L': 12,  # 医师
            'M': 20,  # 医院
            'N': 20,  # 备注
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 写入数据行
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, prescription in enumerate(prescriptions, 2):
            # 序号
            cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # 患者姓名
            cell = ws.cell(row=row_idx, column=2, value=prescription.get('patient_name', ''))
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # 年龄
            cell = ws.cell(row=row_idx, column=3, value=prescription.get('patient_age', ''))
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # 性别
            cell = ws.cell(row=row_idx, column=4, value=prescription.get('patient_gender', ''))
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # 日期
            cell = ws.cell(row=row_idx, column=5, value=prescription.get('date', ''))
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # 症状
            cell = ws.cell(row=row_idx, column=6, value=prescription.get('symptoms', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 诊断
            cell = ws.cell(row=row_idx, column=7, value=prescription.get('diagnosis', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 方剂名称
            cell = ws.cell(row=row_idx, column=8, value=prescription.get('formula_name', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 药材组成
            cell = ws.cell(row=row_idx, column=9, value=prescription.get('herbs', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 剂量
            cell = ws.cell(row=row_idx, column=10, value=prescription.get('dosage', ''))
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # 用法
            cell = ws.cell(row=row_idx, column=11, value=prescription.get('usage', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 医师
            cell = ws.cell(row=row_idx, column=12, value=prescription.get('doctor_name', ''))
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # 医院
            cell = ws.cell(row=row_idx, column=13, value=prescription.get('hospital', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 备注
            cell = ws.cell(row=row_idx, column=14, value=prescription.get('notes', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 设置行高
            ws.row_dimensions[row_idx].height = 30
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存文件
        wb.save(output_path)
        
        return output_path
    
    def export_with_template(self, prescriptions, template_path, output_path):
        """
        使用模板导出处方数据
        
        Args:
            prescriptions: 处方数据列表
            template_path: 模板文件路径
            output_path: 输出文件路径
        """
        from openpyxl import load_workbook
        
        # 加载模板
        wb = load_workbook(template_path)
        ws = wb.active
        
        # 从第2行开始写入数据（假设第1行是标题）
        for row_idx, prescription in enumerate(prescriptions, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=prescription.get('patient_name', ''))
            ws.cell(row=row_idx, column=3, value=prescription.get('patient_age', ''))
            ws.cell(row=row_idx, column=4, value=prescription.get('patient_gender', ''))
            ws.cell(row=row_idx, column=5, value=prescription.get('date', ''))
            ws.cell(row=row_idx, column=6, value=prescription.get('symptoms', ''))
            ws.cell(row=row_idx, column=7, value=prescription.get('diagnosis', ''))
            ws.cell(row=row_idx, column=8, value=prescription.get('formula_name', ''))
            ws.cell(row=row_idx, column=9, value=prescription.get('herbs', ''))
            ws.cell(row=row_idx, column=10, value=prescription.get('dosage', ''))
            ws.cell(row=row_idx, column=11, value=prescription.get('usage', ''))
            ws.cell(row=row_idx, column=12, value=prescription.get('doctor_name', ''))
            ws.cell(row=row_idx, column=13, value=prescription.get('hospital', ''))
            ws.cell(row=row_idx, column=14, value=prescription.get('notes', ''))
        
        # 保存文件
        wb.save(output_path)
        
        return output_path
    
    def export_statistics(self, stats_data, output_path):
        """
        导出统计数据
        
        Args:
            stats_data: 统计数据字典
            output_path: 输出文件路径
        """
        wb = Workbook()
        
        # 概览sheet
        ws_overview = wb.active
        ws_overview.title = '概览'
        
        # 写入概览数据
        overview_headers = ['统计项', '数值']
        for col, header in enumerate(overview_headers, 1):
            cell = ws_overview.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        overview_data = [
            ['总处方数', stats_data.get('total', 0)],
            ['患者数', stats_data.get('patients', 0)],
            ['方剂种类', stats_data.get('formulas', 0)],
            ['本月新增', stats_data.get('monthly', 0)],
        ]
        
        for row_idx, (key, value) in enumerate(overview_data, 2):
            ws_overview.cell(row=row_idx, column=1, value=key)
            ws_overview.cell(row=row_idx, column=2, value=value)
        
        # 药材统计sheet
        ws_herbs = wb.create_sheet('药材统计')
        herb_stats = stats_data.get('herb_stats', {})
        
        ws_herbs.cell(row=1, column=1, value='药材名称')
        ws_herbs.cell(row=1, column=2, value='使用次数')
        
        for row_idx, (herb, count) in enumerate(sorted(herb_stats.items(), key=lambda x: x[1], reverse=True), 2):
            ws_herbs.cell(row=row_idx, column=1, value=herb)
            ws_herbs.cell(row=row_idx, column=2, value=count)
        
        # 方剂统计sheet
        ws_formulas = wb.create_sheet('方剂统计')
        formula_stats = stats_data.get('formula_stats', {})
        
        ws_formulas.cell(row=1, column=1, value='方剂名称')
        ws_formulas.cell(row=1, column=2, value='使用次数')
        
        for row_idx, (formula, count) in enumerate(sorted(formula_stats.items(), key=lambda x: x[1], reverse=True), 2):
            ws_formulas.cell(row=row_idx, column=1, value=formula)
            ws_formulas.cell(row=row_idx, column=2, value=count)
        
        # 月度趋势sheet
        ws_trend = wb.create_sheet('月度趋势')
        trend_stats = stats_data.get('trend_stats', {})
        
        ws_trend.cell(row=1, column=1, value='月份')
        ws_trend.cell(row=1, column=2, value='处方数')
        
        for row_idx, (month, count) in enumerate(sorted(trend_stats.items()), 2):
            ws_trend.cell(row=row_idx, column=1, value=month)
            ws_trend.cell(row=row_idx, column=2, value=count)
        
        # 保存文件
        wb.save(output_path)
        
        return output_path
    
    def create_template(self, output_path):
        """
        创建Excel模板
        
        Args:
            output_path: 输出文件路径
        """
        wb = Workbook()
        ws = wb.active
        ws.title = '处方模板'
        
        # 设置标题行
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置列宽
        column_widths = {
            'A': 8, 'B': 12, 'C': 8, 'D': 8, 'E': 12,
            'F': 30, 'G': 25, 'H': 20, 'I': 40, 'J': 10,
            'K': 25, 'L': 12, 'M': 20, 'N': 20
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 添加示例数据
        example_data = [
            1, '张三', '45', '男', '2024-01-15',
            '头痛、眩晕、失眠多梦、腰膝酸软',
            '肾阴虚，肝阳上亢',
            '六味地黄丸加减',
            '熟地黄 24g，山茱萸 12g，山药 12g，泽泻 9g，茯苓 9g，丹皮 9g',
            '7剂',
            '水煎服，每日一剂，早晚分服',
            '李医生',
            '中医院',
            ''
        ]
        
        for col, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # 保存文件
        wb.save(output_path)
        
        return output_path
